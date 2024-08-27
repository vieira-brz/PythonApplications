import openpyxl
from pathlib import Path

# Define ROOT_DIR como o diretório onde o script está sendo executado
ROOT_DIR = Path(__file__).parent.resolve()

# Cria um novo objeto Workbook (uma nova planilha)
book = openpyxl.Workbook()

# Imprime os nomes das planilhas existentes (por padrão, há uma planilha chamada "Sheet")
print(book.sheetnames)

# Cria uma nova planilha com o nome 'Nome_planilha'
book.create_sheet('Nome_planilha')

# Seleciona a planilha recém-criada para manipulação
book_page = book['Nome_planilha']

# Adiciona uma linha de cabeçalho à planilha
book_page.append(['Coluna 1', 'Coluna 2'])

# Adiciona linhas de dados à planilha
book_page.append(['item 1', 'item 12'])
book_page.append(['item 2', 'item 24'])
book_page.append(['item 3', 'item 36'])

# Salva o arquivo da planilha com o nome 'Planilha Criada.xlsx'
book.save(ROOT_DIR / 'Planilha Criada.xlsx')