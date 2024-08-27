from pathlib import Path
from openpyxl import load_workbook

# Define ROOT_DIR como o diretório onde o script está sendo executado
ROOT_DIR = Path(__file__).parent.resolve()

# Carrega a planilha existente 'Alunos.xlsx'
arquivo = load_workbook(ROOT_DIR / 'Alunos.xlsx')

# Exibe os nomes de todas as abas da planilha
print(arquivo.sheetnames)

# Seleciona a aba ativa (a primeira que aparece ao abrir o arquivo)
aba_atual = arquivo.active
print(aba_atual)

# Seleciona uma aba específica pelo nome
aba_alunos = arquivo['Planilha1']
print(aba_alunos)

# Seleciona e imprime o valor da célula A1
print(aba_alunos['A1'].value)

# Seleciona e imprime o valor da célula na primeira linha e segunda coluna (B1)
print(aba_alunos.cell(row=1, column=2).value)

# Edita o valor da célula na primeira linha e segunda coluna (B1)
aba_alunos.cell(row=1, column=2).value = 'Prova 1'

# Salva as alterações no arquivo 'Alunos.xlsx'
arquivo.save(ROOT_DIR / 'Alunos.xlsx')

# Imprime o número total de linhas e colunas utilizadas na aba
print(aba_alunos.max_row, aba_alunos.max_column)

# Imprime o número de células preenchidas na coluna A
print(len(aba_alunos['A']))