from pathlib import Path
from copy import copy
from openpyxl import load_workbook

ROOT_DIR = Path(__file__).parent.resolve()

arquivo = load_workbook(ROOT_DIR / 'Bairros.xlsx')

aba_basedados = arquivo['Base de Dados']
ultima_linha = aba_basedados.max_row


def criar_aba(bairro, arquivo_bairros, estilos_cabecalho):
    if bairro not in arquivo_bairros.sheetnames:
        arquivo_bairros.create_sheet(bairro)
        nova_aba = arquivo_bairros[bairro]
        nova_aba['A1'].value = 'Data de Nascimento' 
        nova_aba['B1'].value = 'Pessoa' 
        nova_aba['C1'].value = 'Bairro' 
        nova_aba['A1']._style = estilos_cabecalho
        nova_aba['B1']._style = estilos_cabecalho
        nova_aba['C1']._style = estilos_cabecalho
        

def transferir_dados(origem, destino, linha_origem):
    linha_destino = destino.max_row + 1
    for coluna in range(1, 4):
        celula_origem = origem.cell(row=linha_origem, column=coluna)
        celula_destino = destino.cell(row=linha_destino, column=coluna)
        celula_destino.value = celula_origem.value
        celula_destino._style = copy(celula_origem._style)



estilos_cabecalho = copy(aba_basedados['A1']._style)

for linha in range(2, ultima_linha + 1):
    bairro = aba_basedados[f'C{linha}'].value

    if not bairro:
        break

    # Criar aba para o bairro
    criar_aba(bairro, arquivo, estilos_cabecalho)

    # Transferir informações para aba correspondente
    aba_destino = arquivo[bairro]
    transferir_dados(aba_basedados, aba_destino, linha)


arquivo.save(ROOT_DIR / 'Bairros.xlsx')